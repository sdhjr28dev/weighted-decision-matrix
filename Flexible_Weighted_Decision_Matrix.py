import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================
# 🔧 CONFIGURABLE SETTINGS
# ========================

criteria = ["Cost", "Quality", "Delivery Time", "Customer Support"]
options = ["Option A", "Option B", "Option C"]
evaluator_names = ["Evaluator 1", "Evaluator 2"]
filename = "Projectx-decision-matrix.xlsx"



#=========================
# Calculate equal weights
#=========================
equal_weight = round(1.0 / len(criteria), 4)
weights = [equal_weight] * len(criteria)

# ============================
# 📄 CREATE SHEETS
# ============================

with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Main", index=False)

    for evaluator_name in evaluator_names:
        zero_matrix = [[0 for _ in options] for _ in criteria]
        df = pd.DataFrame(zero_matrix, columns=options)
        df.insert(0, "Criteria", criteria)
        df["Notes"] = ""
        df.to_excel(writer, sheet_name=evaluator_name, index=False)

# ============================
# ✍️ STYLE HELPERS
# ============================

bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")
wrap_align = Alignment(wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# ============================
# 📐 FORMAT EVALUATOR SHEETS
# ============================

wb = openpyxl.load_workbook(filename)

for evaluator in evaluator_names:
    ws = wb[evaluator]
    max_row = 1 + len(criteria)
    max_col = ws.max_column

    # Format header row
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # Format data cells
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            header = ws.cell(row=1, column=col).value
            if header == "Notes":
                cell.alignment = wrap_align
            elif header in options:
                cell.alignment = center_align

    # Format Notes column
    notes_col_letter = get_column_letter(max_col)
    ws.column_dimensions[notes_col_letter].width = 40

# ============================
# 📊 MAIN SHEET SETUP
# ============================

main_ws = wb["Main"]

# Move "Main" to the top and set active
wb._sheets.sort(key=lambda ws: ws.title != "Main")
wb.active = 0

headers = ["Criteria"] + options + ["Weight"]
for col_index, header in enumerate(headers, start=1):
    cell = main_ws.cell(row=1, column=col_index, value=header)
    cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    cell.font = bold_font
    cell.alignment = center_align
    cell.border = thin_border

# Fill in criteria and weights
for i, criterion in enumerate(criteria, start=2):
    main_ws.cell(row=i, column=1, value=criterion)
    main_ws.cell(row=i, column=len(options)+2, value=weights[i-2])
    for col in range(1, len(options) + 3):
        main_ws.cell(row=i, column=col).border = thin_border
        if col != 1:
            main_ws.cell(row=i, column=col).alignment = center_align

# Add AVERAGE formulas
for row in range(2, 2 + len(criteria)):
    for opt_idx in range(len(options)):
        col_letter = get_column_letter(opt_idx + 2)
        cell_refs = [f"'{evaluator}'!{col_letter}{row}" for evaluator in evaluator_names]
        avg_formula = f"=AVERAGE({', '.join(cell_refs)})"
        cell = main_ws.cell(row=row, column=opt_idx + 2, value=avg_formula)
        cell.border = thin_border
        cell.alignment = center_align

# Add Total Score row
total_score_row = 2 + len(criteria) + 1
main_ws.cell(row=total_score_row, column=1, value="Total Score").font = bold_font

for opt_idx in range(len(options)):
    col_letter = get_column_letter(opt_idx + 2)
    formula = (
        f"=SUMPRODUCT({col_letter}2:{col_letter}{1+len(criteria)}, "
        f"${get_column_letter(len(options)+2)}$2:${get_column_letter(len(options)+2)}${1+len(criteria)})"
    )
    cell = main_ws.cell(row=total_score_row, column=opt_idx + 2, value=formula)
    cell.number_format = "0.00"
    cell.font = bold_font
    cell.border = thin_border
    cell.alignment = center_align

# ============================
# 🚀 COLUMN WIDTH ADJUSTMENTS
# ============================

def adjust_column_a_width(ws):
    max_length = 0
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
                cell.alignment = wrap_align  # Apply wrap to all Column A
    ws.column_dimensions["A"].width = min(max_length + 2, 40)

def adjust_options_columns_width_from_headers(ws, options):
    max_length = max(len(option) for option in options)
    adjusted_width = max_length + 2
    for idx in range(len(options)):
        col_letter = get_column_letter(idx + 2)  # Options start at column B
        ws.column_dimensions[col_letter].width = adjusted_width

# Apply to evaluator sheets
for evaluator in evaluator_names:
    ws = wb[evaluator]
    adjust_column_a_width(ws)
    adjust_options_columns_width_from_headers(ws, options)

# Apply to main sheet
adjust_column_a_width(main_ws)
adjust_options_columns_width_from_headers(main_ws, options)

# ============================
# 💾 SAVE
# ============================

wb.save(filename)
print(f"✅ File saved as: {filename}")
